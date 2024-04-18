import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import sys
import os

def search_inaturalist(scientific_name):
    url = f"https://www.inaturalist.org/search?q={scientific_name.replace(' ', '+')}"
    print(f"searching inaturalist for {scientific_name}...")
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    
    first_result = soup.select_one('.media-body h4.media-heading a')
    if first_result:
        print(f"got {first_result['href']}!")
        return "https://www.inaturalist.org" + first_result['href']
    else:
        print(f"no result :(")
        return ''

def search_gobotany(scientific_name):
    url = f"https://gobotany.nativeplanttrust.org/search/?q={scientific_name.replace(' ', '+')}"
    print(f"searching gobotony for {scientific_name}...")
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    
    first_result = soup.select_one('#search-results-list li a')
    if first_result:
        print(f"got {first_result['href']}!")
        return "https://gobotany.nativeplanttrust.org" + first_result['href']
    else:
        print(f"no result :(")
        return ''

def process_xlsx(input_file, output_file):
    workbook = load_workbook(filename=input_file)
    sheet = workbook.active
    
    # Determine the columns to use for URL results
    inaturalist_column = sheet.max_column + 1
    gobotany_column = sheet.max_column + 2
    
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        scientific_name = row[0]
        common_name = row[1]
        inaturalist_url = search_inaturalist(scientific_name)
        gobotany_url = search_gobotany(scientific_name)
        sheet.cell(row=row_num, column=inaturalist_column, value=inaturalist_url)
        sheet.cell(row=row_num, column=gobotany_column, value=gobotany_url)
    
    print(f"done!")
    workbook.save(output_file)

# Check if the required command-line arguments are provided
if len(sys.argv) != 3:
    print("Usage: python script.py <input_file> <output_file>")
    sys.exit(1)

# Get the input and output filenames from the command-line arguments
input_filename = sys.argv[1]
output_filename = sys.argv[2]

# Get the absolute file paths
script_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(script_dir, input_filename)
output_file = os.path.join(script_dir, output_filename)

process_xlsx(input_file, output_file)
